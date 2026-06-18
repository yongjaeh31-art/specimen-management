from io import BytesIO
from urllib.parse import quote

import xlrd
from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from fastapi.responses import StreamingResponse
from openpyxl import Workbook, load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment as XAlign, Font as XFont
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.database import get_db
from app.services.import_service import import_orders

router = APIRouter(prefix="/api/import", tags=["imports"])

_TARGET_SPECIMEN = "Urine (Random)"
_TARGET_TEST     = "Ordinary culture & Sensitivity (MIC)"
_HOSP_HEADERS    = {"병원명", "거래처명", "의뢰기관", "의뢰처", "병원"}


@router.post("/orders")
async def import_order_file(
    file: UploadFile = File(...),
    batch_type: str = Form("1차"),
    is_final: bool = Form(False),
    db: Session = Depends(get_db),
):
    try:
        content = await file.read()
        return import_orders(db, file.filename or "orders", content, batch_type, is_final)
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/urine-mark")
async def urine_mark_list(file: UploadFile = File(...)):
    """
    접수리스트 Excel 업로드 →
    검체명=Urine(Random) & 검사명=OC&S 해당 행:
      · 검사명 셀 앞에 ▣ 삽입 (텍스트 직접 삽입)
      · 성명(성별/나이)~병원명 사이 gap 셀에 ● 삽입
    수정된 Excel 반환 (xlsx / xls 모두 지원, 출력은 xlsx)
    """
    content = await file.read()
    filename = file.filename or "orders.xlsx"
    suffix = filename.lower().rsplit(".", 1)[-1]

    if suffix == "xlsx":
        out_bytes, count = _mark_xlsx(content)
        out_name = filename
    elif suffix == "xls":
        out_bytes, count = _mark_xls(content)
        out_name = filename.rsplit(".", 1)[0] + ".xlsx"
    else:
        raise HTTPException(status_code=400, detail="xls 또는 xlsx 파일만 지원합니다.")

    return StreamingResponse(
        BytesIO(out_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(out_name)}",
            "X-Marked-Count": str(count),
        },
    )


# ── 내부 헬퍼 ─────────────────────────────────────────────────────────────────

def _find_target_cols(header_vals: list[str]) -> tuple[list[int], int | None, int | None]:
    """헤더 리스트에서 검사명(들) / 검체명 / 병원명 컬럼 0-indexed 반환"""
    test_cols, spec_col, hosp_col = [], None, None
    for i, v in enumerate(header_vals):
        v = v.strip()
        if v == "검체명":
            spec_col = i
        if v == "검사명" or (v.startswith("검사명") and (len(v) == 3 or v[3:].isdigit())):
            test_cols.append(i)
        if v in _HOSP_HEADERS:
            hosp_col = i
    return test_cols, spec_col, hosp_col


def _insert_symbol_xlsx(ws, row_no: int, col_1: int, symbol: str) -> bool:
    """xlsx 셀에 기호 삽입 (병합 셀이면 분리 후 삽입). 성공 시 True."""
    cell = ws.cell(row_no, col_1)
    if isinstance(cell, MergedCell):
        for mr in list(ws.merged_cells.ranges):
            if (mr.min_row <= row_no <= mr.max_row and
                    mr.min_col <= col_1 <= mr.max_col):
                ws.unmerge_cells(str(mr))
                if mr.min_col < col_1:
                    ws.merge_cells(start_row=mr.min_row, start_column=mr.min_col,
                                   end_row=mr.max_row,   end_column=col_1 - 1)
                if col_1 < mr.max_col:
                    ws.merge_cells(start_row=mr.min_row, start_column=col_1 + 1,
                                   end_row=mr.max_row,   end_column=mr.max_col)
                break
    tgt = ws.cell(row_no, col_1)
    if not isinstance(tgt, MergedCell):
        tgt.value = symbol
        tgt.alignment = XAlign(horizontal="center", vertical="center")
        return True
    return False


def _mark_xlsx(content: bytes) -> tuple[bytes, int]:
    """xlsx: ▣ (검사명 앞) + ● (병원명 앞) 삽입"""
    wb = load_workbook(BytesIO(content))
    total = 0

    for ws in wb.worksheets:
        hdr_row = spec_col = hosp_col_1 = None
        test_cols_1: list[int] = []   # 1-indexed

        for row in ws.iter_rows():
            vals = [str(c.value or "").strip() for c in row]
            tc, sc, hc = _find_target_cols(vals)
            if sc is not None:
                test_cols_1 = [i + 1 for i in tc]
                spec_col    = sc + 1
                hosp_col_1  = (hc + 1) if hc is not None else None
                hdr_row     = row[0].row
                break

        if not hdr_row or spec_col is None or not test_cols_1:
            continue

        between_hosp_col_1 = (   # ● 삽입 위치 (병원명 바로 앞)
            hosp_col_1 - 1 if (hosp_col_1 is not None and hosp_col_1 > 1) else None
        )

        # 매칭 행 수집 (1-indexed)
        mark_rows_1: set[int] = set()
        for row in ws.iter_rows(min_row=hdr_row + 1):
            try:
                spec_val = str(row[spec_col - 1].value or "").strip()
            except IndexError:
                continue
            if spec_val != _TARGET_SPECIMEN:
                continue
            for tc in test_cols_1:
                try:
                    cell = row[tc - 1]
                except IndexError:
                    continue
                if str(cell.value or "").strip() == _TARGET_TEST:
                    mark_rows_1.add(row[0].row)
                    break

        # ● 삽입 (병원명 바로 앞)
        for mr in mark_rows_1:
            if between_hosp_col_1 is not None:
                if _insert_symbol_xlsx(ws, mr, between_hosp_col_1, "●"):
                    total += 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), total


def _mark_xls(content: bytes) -> tuple[bytes, int]:
    """xls 읽기 → openpyxl xlsx 재구성 (폰트·너비·높이·정렬·병합 보존)
    병원명~검사명 사이 셀(between_col)에 ▣ 삽입, 해당 병합 범위 자동 조정"""
    H_ALIGN = {1: "left", 2: "center", 3: "right", 5: "justify"}
    V_ALIGN = {0: "top", 1: "center", 2: "bottom", 3: "justify"}

    wb_r = xlrd.open_workbook(file_contents=content, formatting_info=True)
    wb_w = Workbook()
    wb_w.remove(wb_w.active)
    total = 0

    for si in range(wb_r.nsheets):
        ws_r = wb_r.sheet_by_index(si)
        ws_w = wb_w.create_sheet(ws_r.name or f"Sheet{si + 1}")

        # ① 헤더 + 대상 컬럼 탐색 (mark_rows 계산 전 필수)
        hdr_row_idx = spec_col_idx = hosp_col_idx = None
        test_col_indices: list[int] = []

        for r in range(min(ws_r.nrows, 30)):
            vals = []
            for c in range(ws_r.ncols):
                v = str(ws_r.cell_value(r, c)).strip()
                try:
                    v = v.encode("latin1").decode("cp949")
                except Exception:
                    pass
                vals.append(v)
            tc, sc, hc = _find_target_cols(vals)
            if sc is not None:
                test_col_indices = tc
                spec_col_idx     = sc
                hosp_col_idx     = hc
                hdr_row_idx      = r
                break

        # between_hosp_col: 병원명 바로 앞 컬럼 (0-indexed) → ●
        between_hosp_col: int | None = (
            hosp_col_idx - 1 if (hosp_col_idx is not None and hosp_col_idx > 0) else None
        )

        # ② mark_rows 미리 계산 (0-indexed)
        mark_rows: set[int] = set()
        if hdr_row_idx is not None and spec_col_idx is not None and test_col_indices:
            for r in range(hdr_row_idx + 1, ws_r.nrows):
                if str(ws_r.cell_value(r, spec_col_idx)).strip() == _TARGET_SPECIMEN:
                    for tc in test_col_indices:
                        if str(ws_r.cell_value(r, tc)).strip() == _TARGET_TEST:
                            mark_rows.add(r)
                            break

        # ③ 컬럼 너비
        for c in range(ws_r.ncols):
            ci = ws_r.colinfo_map.get(c)
            if ci and ci.width > 0:
                ws_w.column_dimensions[get_column_letter(c + 1)].width = round(ci.width / 256, 1)

        # 기호 삽입 대상 컬럼 집합 (병합 분리 필요 여부 판단에 사용)
        split_cols = {c for c in (between_hosp_col,) if c is not None}

        # ④ 병합 셀 복사
        #    mark_rows 포함 병합에서 split_cols 중 하나가 마지막 컬럼이면 분리
        for rlo, rhi, clo, chi in ws_r.merged_cells:
            if rhi <= rlo or chi <= clo:
                continue
            last_c = chi - 1   # 0-indexed 마지막 컬럼
            needs_split = (
                last_c in split_cols
                and any(r in mark_rows for r in range(rlo, rhi))
            )
            try:
                if needs_split and last_c > clo:
                    ws_w.merge_cells(
                        start_row=rlo + 1, start_column=clo + 1,
                        end_row=rhi,       end_column=chi - 1,
                    )
                elif not needs_split:
                    ws_w.merge_cells(
                        start_row=rlo + 1, start_column=clo + 1,
                        end_row=rhi,       end_column=chi,
                    )
            except Exception:
                pass

        # ⑤ 셀 복사 + ● 삽입 (병원명 바로 앞)
        for r in range(ws_r.nrows):
            ri = ws_r.rowinfo_map.get(r)
            if ri and ri.height > 0:
                ws_w.row_dimensions[r + 1].height = round(ri.height / 20, 1)

            is_mark = r in mark_rows

            for c in range(ws_r.ncols):
                ctype = ws_r.cell_type(r, c)
                val   = ws_r.cell_value(r, c)

                # EMPTY(0) 또는 BLANK(6): 비앵커 병합 셀은 xlrd에서 BLANK로 반환
                if ctype in (xlrd.XL_CELL_EMPTY, xlrd.XL_CELL_BLANK):
                    if is_mark and c == between_hosp_col:
                        tgt = ws_w.cell(r + 1, c + 1)
                        if not isinstance(tgt, MergedCell):
                            tgt.value = "●"
                            tgt.alignment = XAlign(horizontal="center", vertical="center")
                            total += 1
                    continue

                if ctype == xlrd.XL_CELL_NUMBER:
                    if isinstance(val, float) and val.is_integer():
                        val = int(val)
                elif ctype == xlrd.XL_CELL_DATE:
                    try:
                        val = xlrd.xldate_as_datetime(val, wb_r.datemode)
                    except Exception:
                        pass

                target = ws_w.cell(r + 1, c + 1)
                if isinstance(target, MergedCell):
                    continue
                target.value = val
                oc = target

                xf  = wb_r.xf_list[ws_r.cell_xf_index(r, c)]
                fnt = wb_r.font_list[xf.font_index]
                try:
                    fn = fnt.name.encode("latin1").decode("cp949")
                except Exception:
                    fn = fnt.name or "굴림"
                oc.font = XFont(name=fn, size=round(fnt.height / 20, 1), bold=bool(fnt.bold))
                oc.alignment = XAlign(
                    horizontal=H_ALIGN.get(xf.alignment.hor_align),
                    vertical=V_ALIGN.get(xf.alignment.vert_align, "bottom"),
                    wrap_text=bool(xf.alignment.text_wrapped),
                )

    out = BytesIO()
    wb_w.save(out)
    return out.getvalue(), total
