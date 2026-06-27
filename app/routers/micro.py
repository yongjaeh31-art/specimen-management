from datetime import date, timedelta
from io import BytesIO
from math import ceil

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from sqlalchemy.orm import Session

from app.database import get_db
from app.models import DepartmentSubcategoryAssignment, MicroCultureAssignment, MicroCulturePlan
from app.schemas import (
    AutoCultureScanRequest, CultureScanRequest,
    DEPARTMENT_SUBCATEGORIES, MICRO_CULTURE_TYPES, SubdivisionScanRequest,
)
from app.services.micro_service import (
    _assignment_dict, _shift_start, _today_start,
    assign_culture_hole,
    auto_assign_culture_hole,
    get_culture_plan_list,
    get_today_micro_assignments,
)
from app.services.subdivision_service import assign_department_subcategory, combined_assignments

router = APIRouter(prefix="/api/micro", tags=["micro"])


@router.post("/culture-scan")
def culture_scan(request: CultureScanRequest, db: Session = Depends(get_db)):
    try:
        return assign_culture_hole(
            db,
            request.accession_no,
            request.culture_type,
            request.client_name,
            request.operator_name,
            request.workstation_name,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/auto-culture-scan")
def auto_culture_scan(request: AutoCultureScanRequest, db: Session = Depends(get_db)):
    """culture_type 선택 없이 자동 판정 → 50칸 랙 위치 배정"""
    try:
        return auto_assign_culture_hole(
            db,
            request.accession_no,
            request.client_name,
            request.operator_name,
            request.workstation_name,
            rack_size=request.rack_size,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/today-assignments")
def today_assignments(db: Session = Depends(get_db)):
    """오늘 처리된 전체 미생물 소분류 작업 목록"""
    return get_today_micro_assignments(db)


@router.get("/culture-summary")
def culture_summary(db: Session = Depends(get_db)):
    """모든 culture_type의 오늘 진행 현황 요약 (왼쪽 패널 배지용)"""
    from sqlalchemy import func as sqlfunc
    today = _today_start()
    rows = (
        db.query(
            MicroCulturePlan.culture_type,
            MicroCulturePlan.status,
            sqlfunc.count(MicroCulturePlan.id).label("cnt"),
        )
        .filter(MicroCulturePlan.planned_at >= today)
        .group_by(MicroCulturePlan.culture_type, MicroCulturePlan.status)
        .all()
    )
    summary: dict[str, dict] = {ct: {"culture_type": ct, "total": 0, "done": 0, "pending": 0} for ct in MICRO_CULTURE_TYPES}
    for ct, status, cnt in rows:
        if ct in summary:
            summary[ct]["total"] += cnt
            if status == "DONE":
                summary[ct]["done"] += cnt
            else:
                summary[ct]["pending"] += cnt
    return [v for v in summary.values() if v["total"] > 0]


@router.get("/culture-types")
def culture_types():
    return MICRO_CULTURE_TYPES


@router.get("/culture-plans")
def culture_plans(culture_type: str, full: bool = False, db: Session = Depends(get_db)):
    """선택한 culture_type의 예정 목록(PENDING) + 완료 목록(DONE) 조회
    full=False(기본): 남은 목록 최대 30건만 반환 (주기 새로고침용 경량 응답)
    full=True: 남은 목록 전체 반환 (펼쳐보기용)"""
    return get_culture_plan_list(db, culture_type, full=full)


@router.get("/culture-assignments")
def culture_assignments(db: Session = Depends(get_db)):
    rows = (
        db.query(MicroCultureAssignment)
        .order_by(MicroCultureAssignment.assigned_at.desc())
        .limit(500)
        .all()
    )
    return [_assignment_dict(row) for row in rows]


@router.get("/subcategories")
def subcategories():
    return DEPARTMENT_SUBCATEGORIES


@router.post("/subcategory-scan")
def subcategory_scan(request: SubdivisionScanRequest, db: Session = Depends(get_db)):
    try:
        return assign_department_subcategory(
            db,
            request.accession_no,
            request.department_name,
            request.subcategory,
            request.client_name,
            request.operator_name,
            request.workstation_name,
        )
    except ValueError as exc:
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.get("/subcategory-assignments")
def subcategory_assignments(db: Session = Depends(get_db)):
    return combined_assignments(db)


@router.post("/reset-shift")
def reset_shift(db: Session = Depends(get_db)):
    """출근 기준(KST 19:00) 이후 소분류 스캔 전체 초기화"""
    since = _shift_start()

    # 미생물 배정 삭제 + plan PENDING 복원
    micro_list = (
        db.query(MicroCultureAssignment)
        .filter(MicroCultureAssignment.assigned_at >= since)
        .all()
    )
    for asn in micro_list:
        plan = (
            db.query(MicroCulturePlan)
            .filter(
                MicroCulturePlan.accession_no == asn.accession_no,
                MicroCulturePlan.culture_type == asn.culture_type,
                MicroCulturePlan.assignment_id == asn.id,
            )
            .first()
        )
        if plan:
            plan.status = "PENDING"
            plan.assignment_id = None
        db.delete(asn)

    # 비미생물 소분류 배정 삭제
    dept_count = (
        db.query(DepartmentSubcategoryAssignment)
        .filter(DepartmentSubcategoryAssignment.assigned_at >= since)
        .delete(synchronize_session=False)
    )

    db.commit()
    kst = since + timedelta(hours=9)
    return {
        "status": "OK",
        "shift_start_kst": kst.strftime("%Y-%m-%d %H:%M"),
        "deleted_micro": len(micro_list),
        "deleted_dept": dept_count,
    }


@router.get("/worklist/export")
def export_worklist(db: Session = Depends(get_db)):
    """오늘 스캔 완료 목록 → {소분류} 워크리스트 (첨부 서식 그대로, 25행/페이지, LIS순)"""
    from openpyxl.styles import Color as XLColor

    assignments = get_today_micro_assignments(db)
    fallback_date = date.today().strftime("%Y-%m-%d")

    # 소분류별 그룹핑 → LIS 순번(culture_order) 오름차순
    groups: dict[str, list] = {}
    for row in assignments:
        groups.setdefault(row["culture_type"], []).append(row)
    for ct in groups:
        groups[ct].sort(key=lambda r: r.get("culture_order") or 0)

    wb = Workbook()
    wb.remove(wb.active)

    # ── 참조 xlsx 서식 그대로 재현 ────────────────────────────────────────
    # 배경 fill: Excel 기본 흰 배경 (theme=0 / bgColor indexed=64)
    auto_fill = PatternFill(
        patternType="solid",
        fgColor=XLColor(theme=0, tint=0.0),
        bgColor=XLColor(indexed=64),
    )

    # 제목 하단 선 (#999999)
    title_border = Border(bottom=Side(border_style="thin", color="FF999999"))

    # 헤더: 사방 thin
    hdr_border = Border(
        top=Side(border_style="thin"),
        bottom=Side(border_style="thin"),
        left=Side(border_style="thin"),
        right=Side(border_style="thin"),
    )

    def _data_border(is_first: bool, ci: int) -> Border:
        """col 1:left thin, col 6:right thin / 첫째 데이터행 top thin, 이후 hair(#BBB)"""
        hair = Side(border_style="hair", color="FFBBBBBB")
        top  = Side(border_style="thin") if is_first else hair
        lft  = Side(border_style="thin") if ci == 1 else None
        rgt  = Side(border_style="thin") if ci == 6 else None
        return Border(top=top, bottom=hair, left=lft, right=rgt)

    COL_NAMES  = ["순번", "접수번호", "성명 (성별/나이)", "병원명", "검사명", "검체명"]
    COL_WIDTHS = [5.0, 13.0, 16.0, 20.0, 38.0, 14.0]
    NCOLS = len(COL_NAMES)
    DATA_PER_PAGE = 25
    HEADER_ROWS   = 3  # 타이틀 + 날짜/페이지 + 헤더

    if not groups:
        ws = wb.create_sheet("미생물 워크리스트")
        ws.cell(1, 1, "오늘 처리된 미생물 소분류 검체가 없습니다.")
    else:
        for ct, rows in groups.items():
            ws = wb.create_sheet(_safe_sheet_name(ct)[:31])

            ws.page_setup.paperSize   = 9            # A4
            ws.page_setup.orientation = "landscape"
            ws.page_setup.fitToPage   = True
            ws.page_setup.fitToWidth  = 1
            ws.page_setup.fitToHeight = 0
            ws.page_margins.left   = 0.5
            ws.page_margins.right  = 0.5
            ws.page_margins.top    = 0.7
            ws.page_margins.bottom = 0.7

            for i, w in enumerate(COL_WIDTHS, 1):
                ws.column_dimensions[get_column_letter(i)].width = w

            total_pages = ceil(len(rows) / DATA_PER_PAGE)

            for pg in range(total_pages):
                page_rows = rows[pg * DATA_PER_PAGE : (pg + 1) * DATA_PER_PAGE]
                base = pg * (HEADER_ROWS + DATA_PER_PAGE) + 1

                # ── 행1: "{소분류명} 워크리스트" ─────────────────────────
                r1 = base
                ws.merge_cells(start_row=r1, start_column=1, end_row=r1, end_column=NCOLS)
                tc = ws.cell(r1, 1, f"{ct} 워크리스트")
                tc.font      = Font(name="굴림", size=14, bold=True)
                tc.alignment = Alignment(horizontal="center", vertical="center")
                tc.border    = title_border
                ws.row_dimensions[r1].height = 24.0

                # ── 행2: 접수일자(좌) + 페이지(우) ──────────────────────
                r2 = base + 1
                # 접수일자: 해당 페이지 데이터의 실제 날짜 범위
                dates = [r.get("accession_date") for r in page_rows if r.get("accession_date")]
                date_min = min(dates) if dates else fallback_date
                date_max = max(dates) if dates else fallback_date
                ws.merge_cells(start_row=r2, start_column=1, end_row=r2, end_column=4)
                c = ws.cell(r2, 1, f"접수일자 :   {date_min}  ~  {date_max}")
                c.font      = Font(name="굴림", size=9)
                c.alignment = Alignment(vertical="center")
                ws.merge_cells(start_row=r2, start_column=5, end_row=r2, end_column=NCOLS)
                c = ws.cell(r2, 5, f"페이지 : {pg + 1} of {total_pages}")
                c.font      = Font(name="굴림", size=9)
                c.alignment = Alignment(horizontal="right", vertical="center")
                ws.row_dimensions[r2].height = 14.0

                # ── 행3: 컬럼 헤더 ──────────────────────────────────────
                r3 = base + 2
                for ci, h in enumerate(COL_NAMES, 1):
                    c = ws.cell(r3, ci, h)
                    c.font      = Font(name="굴림", size=10, bold=True)
                    c.fill      = auto_fill
                    c.border    = hdr_border
                    c.alignment = Alignment(horizontal="center", vertical="center")
                ws.row_dimensions[r3].height = 16.05

                # ── 데이터 행 ───────────────────────────────────────────
                for i, row in enumerate(page_rows):
                    dr  = base + HEADER_ROWS + i
                    seq = pg * DATA_PER_PAGE + i + 1
                    vals = [
                        seq,
                        row.get("accession_no", ""),
                        _patient_text(row),
                        row.get("hospital_name") or "",
                        ", ".join(row.get("test_names") or []),
                        row.get("specimen_name") or "",
                    ]
                    for ci, val in enumerate(vals, 1):
                        c = ws.cell(dr, ci, val)
                        c.font      = Font(name="굴림", size=10, bold=(ci == 2))
                        c.fill      = auto_fill
                        c.border    = _data_border(i == 0, ci)
                        c.alignment = Alignment(
                            horizontal="center" if ci <= 2 else "left",
                            vertical="center",
                        )
                    ws.row_dimensions[dr].height = 14.0

                # 페이지 나누기 (마지막 제외)
                if pg < total_pages - 1:
                    ws.row_breaks.append(Break(id=base + HEADER_ROWS + len(page_rows) - 1))

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    fname = f"micro_worklist_{date.today().strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={fname}"},
    )


@router.get("/subcategory-assignments/export")
def export_subcategory_assignments(
    department_name: str | None = None,
    subcategory: str | None = None,
    page_no: int | None = None,
    db: Session = Depends(get_db),
):
    rows = combined_assignments(db, from_date=_shift_start())
    if department_name:
        rows = [row for row in rows if row["department_name"] == department_name]
    if subcategory:
        rows = [row for row in rows if row["subcategory"] == subcategory]
    if page_no:
        rows = [row for row in rows if row["page_no"] == page_no]

    workbook = Workbook()
    workbook.remove(workbook.active)
    fallback_date_str = date.today().strftime("%Y-%m-%d")

    # 25컬럼 너비 (출력예시.xlsx 서식 기준)
    COL_WIDTHS = [4.5625, 9.125, 0.125, 5.0, 5.3125, 1.125, 4.4375, 0.125, 4.125, 2.0,
                  11.6875, 0.875, 0.6875, 0.125, 5.5625, 4.4375, 8.5625, 3.4375, 6.875,
                  6.4375, 4.4375, 1.3125, 11.6875, 2.125, 3.6875]
    # 검체 1건당 6행 구조의 행 높이
    REC_HEIGHTS = [9.75, 0.75, 1.5, 12.75, 1.5, 1.5]
    FONT        = "맑은 고딕"
    dotted_left = Border(left=Side(border_style="dotted"))
    medium_bot  = Side(border_style="medium")

    def _set_border_bot(cell):
        """기존 테두리를 유지하면서 하단 medium 추가"""
        b = cell.border
        cell.border = Border(
            top=b.top, bottom=medium_bot,
            left=b.left, right=b.right,
        )

    def _build_sheet(ws, dept, subcat, grp_rows, _fallback=fallback_date_str):
        ws.page_setup.paperSize  = 9
        ws.page_setup.orientation = "portrait"
        ws.page_margins.left     = 0.5
        ws.page_margins.right    = 0.5
        ws.page_margins.top      = 0.7
        ws.page_margins.bottom   = 0.7
        ws.print_options.horizontalCentered = True   # 수평 가운데 정렬

        for i, w in enumerate(COL_WIDTHS, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        # ── 행1: 타이틀 ───────────────────────────────────────────────
        ws.row_dimensions[1].height = 27.75
        ws.merge_cells("E1:U1")        # cols 5-21 (25열 기준)
        tc = ws.cell(1, 5, "워크리스트")
        tc.font      = Font(name=FONT, size=20, bold=True)
        tc.alignment = Alignment(horizontal="center", vertical="center")

        # ── 행2: 주 헤더 (병합 → 스타일) ─────────────────────────────
        ws.row_dimensions[2].height = 12.75
        ws.merge_cells("C2:E2")        # 수진자명 cols 3-5
        ws.merge_cells("F2:J2")        # 병리번호 cols 6-10
        ws.merge_cells("K2:M2")        # 거래처명 cols 11-13
        ws.merge_cells("N2:X2")        # 검사명 cols 14-24
        for col, label in [(1,"순번"),(2,"접수일자"),(3,"수진자명"),
                           (6,"병리번호"),(11,"거래처명"),(14,"검사명"),(25,"건수")]:
            c = ws.cell(2, col, label)
            c.font      = Font(name=FONT, size=8, color="808080")
            c.alignment = Alignment(horizontal="left", vertical="center")

        # ── 행3: 보조 헤더 (병합 → 스타일 → 하단 medium 선) ─────────
        ws.row_dimensions[3].height = 13.5
        ws.merge_cells("C3:E3")
        ws.merge_cells("F3:J3")        # 수탁바코드 cols 6-10
        ws.merge_cells("K3:M3")        # cols 11-13
        ws.merge_cells("N3:X3")        # cols 14-24
        for col, label in [(2,"등록번호"),(3,"생년월일"),(6,"수탁바코드")]:
            c = ws.cell(3, col, label)
            c.font      = Font(name=FONT, size=8, color="808080")
            c.alignment = Alignment(horizontal="left", vertical="center")
        # medium 하단선: 각 병합 구역의 top-left 셀 + 단독 셀에만 적용
        for col in [1, 2, 3, 6, 11, 14, 25]:
            _set_border_bot(ws.cell(3, col))

        # ── 데이터 행 (검체당 6행) ────────────────────────────────────
        for idx, row in enumerate(grp_rows):
            R = 4 + idx * 6

            for ri, h in enumerate(REC_HEIGHTS):
                ws.row_dimensions[R + ri].height = h

            # ① 병합 먼저 (출력예시.xlsx 25열 구조)
            ws.merge_cells(start_row=R,   start_column=1,  end_row=R+3, end_column=1)   # A    순번
            ws.merge_cells(start_row=R,   start_column=2,  end_row=R+1, end_column=3)   # BC   접수일자
            ws.merge_cells(start_row=R,   start_column=4,  end_row=R+2, end_column=5)   # DE   수진자명
            ws.merge_cells(start_row=R,   start_column=6,  end_row=R+1, end_column=10)  # F-J  병리번호
            ws.merge_cells(start_row=R,   start_column=11, end_row=R+3, end_column=14)  # K-N  거래처명
            ws.merge_cells(start_row=R,   start_column=15, end_row=R,   end_column=23)  # O-W  검사명
            ws.merge_cells(start_row=R+1, start_column=15, end_row=R+3, end_column=23)  # O-W  검체명
            ws.merge_cells(start_row=R,   start_column=25, end_row=R+3, end_column=25)  # Y    건수
            ws.merge_cells(start_row=R+3, start_column=2,  end_row=R+4, end_column=3)   # BC   등록번호
            ws.merge_cells(start_row=R+3, start_column=4,  end_row=R+4, end_column=5)   # DE   생년월일
            ws.merge_cells(start_row=R+3, start_column=6,  end_row=R+4, end_column=10)  # F-J  수탁바코드

            seq   = row.get("group_item_no") or (idx + 1)
            pname = row.get("patient_name") or ""
            page  = row.get("patient_age") or ""
            hosp  = row.get("hospital_name") or ""
            tests = ", ".join(row.get("test_names") or [])
            spec  = row.get("specimen_name") or ""
            accno = row.get("accession_no") or ""
            acc_date = row.get("accession_date") or _fallback

            def _c(r, col, val, bold=False, size=8, halign="left", valign="top", border=None):
                cell = ws.cell(r, col, val)
                cell.font      = Font(name=FONT, size=size, bold=bold)
                cell.alignment = Alignment(horizontal=halign, vertical=valign, wrap_text=False)
                if border:
                    cell.border = border
                return cell

            # ② 병합 후 스타일/값 적용 (출력예시.xlsx 25열 구조)
            _c(R,   1,  seq,      halign="left")
            _c(R,   2,  acc_date)
            _c(R,   4,  pname,    bold=True)
            _c(R,   6,  "",       size=8)              # 병리번호 공란
            _c(R,   11, hosp)                          # 거래처명 col 11
            _c(R,   15, tests,    border=dotted_left)  # 검사명 col 15
            _c(R,   25, 1,        halign="right")      # 건수 col 25
            _c(R+1, 15, spec,     size=6, border=dotted_left)  # 검체명 col 15
            # R+2, R+3 col15는 O{R+1}:W{R+3} 병합 보조셀 → 값/테두리 직접 설정 불가
            _c(R+3, 2,  accno)
            _c(R+3, 4,  page,     bold=True)
            _c(R+3, 6,  "",       size=8)              # 수탁바코드 공란
            # 레코드 구분선: R+5(spacer) 상단 = 레코드 블록 하단에 medium 선 전체 적용
            # R+5는 병합이 없는 spacer 행이므로 모든 열에 border 설정 가능
            top_sep = Border(top=medium_bot)
            for _sc in range(1, 26):
                ws.cell(R + 5, _sc).border = top_sep

    if not rows:
        ws = workbook.create_sheet("소분류")
        _build_sheet(ws, "", "소분류 현황", [])
    else:
        groups: dict[tuple, list] = {}
        for row in rows:
            key = (row["department_name"], row["subcategory"])
            groups.setdefault(key, []).append(row)
        for (dept, subcat), grp_rows in groups.items():
            safe_name = _safe_sheet_name(f"{dept}-{subcat}")[:31]
            ws = workbook.create_sheet(safe_name)
            _build_sheet(ws, dept, subcat, grp_rows)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=subdivision_assignments.xlsx"},
    )


def _write_title_row(sheet, title: str, num_cols: int) -> None:
    """1행에 소분류명 타이틀 작성 (셀 병합 + 스타일)"""
    cell = sheet.cell(row=1, column=1, value=title)
    cell.font = Font(bold=True, size=13, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    if num_cols > 1:
        sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    sheet.row_dimensions[1].height = 30


def _format_sheet(sheet, num_cols: int = 8) -> None:
    """2행 헤더 + 3행~ 데이터 스타일링"""
    header_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[2]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    # 처리시간 없이 8컬럼: 번호·워크리스트순·접수번호·성명나이·병원명·검사명·검체명·위치번호
    col_widths = [8, 14, 16, 18, 22, 46, 18, 18]
    for idx, width in enumerate(col_widths[:num_cols], 1):
        sheet.column_dimensions[chr(64 + idx)].width = width
    for row in sheet.iter_rows(min_row=3):
        for cell in row:
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    sheet.freeze_panes = "A3"  # 1행(타이틀)+2행(헤더) 고정


def _safe_sheet_name(value: str) -> str:
    for char in ["\\", "/", "*", "?", ":", "[", "]"]:
        value = value.replace(char, "-")
    return value or "소분류"


def _patient_text(row: dict) -> str:
    name = row.get("patient_name") or ""
    age = row.get("patient_age") or ""
    if name and age:
        return f"{name} / {age}세"
    return name or age
